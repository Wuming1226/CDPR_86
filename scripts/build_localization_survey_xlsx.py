#!/usr/bin/env python3
"""Generate Excel survey from CDPR / visual localization literature summary."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROWS = [
    # --- CDPR 大尺度定位 ---
    {
        "category": "CDPR大尺度定位",
        "name": "仅绳长/卷扬编码器（开环）",
        "method": "绳长正解，少标定或几何标定不足",
        "accuracy": "毫米～厘米；平面大跨度标定后仍有约 ±5～9 mm",
        "link": "https://hal.science/hal-03758216/file/ARK2022_Wang_Cardou_Caro_FV.pdf",
    },
    {
        "category": "CDPR大尺度定位",
        "name": "运动学/几何标定（外测真值）",
        "method": "激光跟踪仪、动捕或视觉作真值，辨识锚点/绳长零点等",
        "accuracy": "标定后静态末端约 ~0.6 mm（3D 打印 CDPR）",
        "link": "https://www.mdpi.com/1424-8220/18/7/2392",
    },
    {
        "category": "CDPR大尺度定位",
        "name": "IPAnema 3 + DCLM 绳长激光",
        "method": "滑轮处激光直接测绳长，控制补偿弹性/蠕变",
        "accuracy": "相对纯编码器：空载精度提升 ~61%，带载 ~86%",
        "link": "https://www.ipa.fraunhofer.de/en/reference_projects/accuracy_improvement_cable-driven_parallel_robots_DCLM-Sensors.html",
    },
    {
        "category": "CDPR大尺度定位",
        "name": "CoGiRo 摄影测量+多边测量+闭环",
        "method": "工业摄影测量与多边测量融合，位姿反馈控制",
        "accuracy": "静态对位 < 250 μm（11×15×6 m，500 kg）",
        "link": "https://scienceportal.tecnalia.com/en/publications/positioning-of-a-cable-driven-parallel-robot-at-better-than-250-%CE%BC/",
    },
    {
        "category": "CDPR大尺度定位",
        "name": "NIST RoboCrane",
        "method": "缆索并联/Stewart 类大尺度原型，实验+仿真",
        "accuracy": "平移约 1 mm，角度约 0.5°（~100 m³ 工作体积）",
        "link": "https://www.nist.gov/programs-projects/robocrane",
    },
    {
        "category": "CDPR大尺度定位",
        "name": "大尺度建造 3D 打印 CDPR",
        "method": "缆索悬挂打印，工作空间约 13.6×9.4×3.3 m",
        "accuracy": "工艺/轨迹：跟踪 <0.4 mm（0.1 m/s）；打印面 0.11～0.7 mm",
        "link": "https://link.springer.com/article/10.1007/s41693-017-0008-0",
    },
    {
        "category": "CDPR大尺度定位",
        "name": "未标定单目测姿（建造 CDPR）",
        "method": "单目相机测末端误差，与激光跟踪仪对比",
        "accuracy": "工作空间中部约 ~18 mm；边缘更大（绳垂）",
        "link": "https://link.springer.com/article/10.1007/s10846-021-01486-z",
    },
    {
        "category": "CDPR大尺度定位",
        "name": "万向节滑轮大尺度 CDPR",
        "method": "机构设计+滑轮/绳变形建模分析精度",
        "accuracy": "取决于设计与张力；文献侧重模型与误差源分析",
        "link": "https://hal.science/hal-03393239v1/file/Accuracy%20of%20large-scale%20cable-driven%20parallel%20robot%20with%20universal%20joint%20pulleys.pdf",
    },
    {
        "category": "CDPR大尺度定位",
        "name": "滑轮运动学标定（专利/方法）",
        "method": "滑轮出索点运动学纳入标定",
        "accuracy": "提高终端控制精度（具体 mm 见原文实验）",
        "link": "https://patents.google.com/patent/CN112518738B/zh",
    },
    # --- CDPR 视觉方案 ---
    {
        "category": "CDPR视觉方案",
        "name": "CoGiRo 工业摄影测量（单系统）",
        "method": "多相机工业摄影测量测平台位姿",
        "accuracy": "坐标不确定度典型 0.2 mm；遮挡时 ~0.5 mm；~2 Hz",
        "link": "https://hal-lirmm.ccsd.cnrs.fr/lirmm-05419284v1/document",
    },
    {
        "category": "CDPR视觉方案",
        "name": "CoGiRo 多边测量（融合后）",
        "method": "摄影测量辅助对齐多边测量站",
        "accuracy": "静态位置 <70 μm，姿态 <110 μrad；闭环后 <250 μm",
        "link": "https://scienceportal.tecnalia.com/en/publications/positioning-of-a-cable-driven-parallel-robot-at-better-than-250-%CE%BC/",
    },
    {
        "category": "CDPR视觉方案",
        "name": "CoGiRo 多相机位姿视觉伺服",
        "method": "多相机+测力；悬链/弹性绳模型；PBVS",
        "accuracy": "最大误差 <1 cm（位置），<0.5°（姿态）；15×11×6 m",
        "link": "https://hal.science/lirmm-02157768/",
    },
    {
        "category": "CDPR视觉方案",
        "name": "建造 CDPR 单目 ArUco",
        "method": "固定三脚架单目+ArUco 板；测重复性与方向精度",
        "accuracy": "中部约 ~18 mm（与 FARO 激光跟踪一致）",
        "link": "https://link.springer.com/article/10.1007/s10846-021-01486-z",
    },
    {
        "category": "CDPR视觉方案",
        "name": "3D 打印 CDPR 动捕标定",
        "method": "Vicon 等光学动捕作真值，绳长残差标定",
        "accuracy": "标定后末端位置误差约 0.6157 mm",
        "link": "https://www.mdpi.com/1424-8220/18/7/2392",
    },
    {
        "category": "CDPR视觉方案",
        "name": "平面仓储 CDPR AprilTag 标定",
        "method": "末端相机观测货架 AprilTag，几何参数辨识",
        "accuracy": "标定后定位 <1 mm",
        "link": "https://www.sciopen.com/article/10.16511/j.cnki.qhdxxb.2022.21.026",
    },
    {
        "category": "CDPR视觉方案",
        "name": "模块化 CDPR 自标定视觉",
        "method": "模块化 CDPR 自标定视觉+LM 优化",
        "accuracy": "文献侧重标定流程；精度见论文实验表",
        "link": "https://www.mdpi.com/1424-8220/26/7/2204",
    },
    {
        "category": "CDPR视觉方案",
        "name": "并联机构 fiducial 绝对位姿（类比）",
        "method": "基座相机+动平台标记阵列直接测姿",
        "accuracy": "亚 mm 级可能（hexapod 等，见原文）",
        "link": "https://www.mdpi.com/1424-8220/22/5/1995",
    },
    {
        "category": "CDPR视觉方案",
        "name": "工业机械臂+摄影测量外环",
        "method": "C-Track 780 等摄影测量增强视觉伺服",
        "accuracy": "静态约 ±0.05 mm / ±0.05°；动态轨迹约 ±0.20 mm / ±0.10°",
        "link": "https://spectrum.library.concordia.ca/id/eprint/993272/",
    },
    {
        "category": "CDPR视觉方案",
        "name": "深度学习视觉伺服（通用，非 CDPR）",
        "method": "Siamese CNN 位姿估计+视觉伺服",
        "accuracy": "平移约 0.6 mm，旋转约 0.4°（特定装配任务）",
        "link": "https://arxiv.org/abs/1903.04713",
    },
    # --- 通用视觉定位 ---
    {
        "category": "通用视觉定位",
        "name": "AprilTag 近距离（实验室）",
        "method": "solvePnP+精确相机标定；与 Vicon 对比",
        "accuracy": "平移均值约 ~0.25 mm",
        "link": "https://www.cdiorio.dev/projects/camera-calibration/",
    },
    {
        "category": "通用视觉定位",
        "name": "AprilTag 工业常用距离",
        "method": "已知 Tag 边长+内参；1～3 m 典型",
        "accuracy": "约 1～3 mm",
        "link": "https://github.com/AprilRobotics/apriltag",
    },
    {
        "category": "通用视觉定位",
        "name": "AprilTag 多算法评估",
        "method": "多种 fiducial 位姿算法实验对比",
        "accuracy": "位置约 1.45 ± 0.82 mm；姿态约 0.26 ± 0.21°",
        "link": "https://link.springer.com/article/10.1007/s42979-024-02993-0",
    },
    {
        "category": "通用视觉定位",
        "name": "ArUco 标记定位",
        "method": "OpenCV ArUco detectMarkers + estimatePoseBoard",
        "accuracy": "随距离与标定变化；中距离常见 mm～cm",
        "link": "https://docs.opencv.org/4.x/d5/dae/tutorial_aruco_detection.html",
    },
    {
        "category": "通用视觉定位",
        "name": "双目/RGB-D 近场",
        "method": "视差/深度图+点云或 ICP",
        "accuracy": "深度噪声常见 1～5 mm；位姿 mm～cm",
        "link": "https://www.intelrealsense.com/",
    },
    {
        "category": "通用视觉定位",
        "name": "光学动捕 Vicon/OptiTrack",
        "method": "多红外相机跟踪反光刚体",
        "accuracy": "位置约 0.1～1 mm；角度约 0.01～0.1°",
        "link": "https://www.vicon.com/",
    },
    {
        "category": "通用视觉定位",
        "name": "工业摄影测量（计量）",
        "method": "多站位拍照+三角化+光束法平差",
        "accuracy": "约 0.05～0.5 mm；大体积仍可达 0.1～1 mm",
        "link": "https://www.mdpi.com/2218-6581/15/5/86",
    },
    {
        "category": "通用视觉定位",
        "name": "Atracsys fusionTrack 500",
        "method": "双相机光学跟踪+被动/主动标记",
        "accuracy": "约 2 m 距离 ~0.08 mm RMS；~335 Hz",
        "link": "https://atracsys.com/product/fusiontrack-500/",
    },
    {
        "category": "通用视觉定位",
        "name": "视觉 SLAM / VIO（ORB-SLAM 等）",
        "method": "特征/直接法+可选 IMU；相对轨迹",
        "accuracy": "室内常 cm 级 drift；长距离 dm～m",
        "link": "https://github.com/UZ-SLAMLab/ORB_SLAM3",
    },
    {
        "category": "通用视觉定位",
        "name": "Fiducial SLAM 对比研究",
        "method": "AprilTag 等辅助 SLAM 与纯特征对比",
        "accuracy": "优于纯特征 SLAM；绝对精度仍依赖地图/Tag 网",
        "link": "https://arxiv.org/abs/2309.04441",
    },
    {
        "category": "通用视觉定位",
        "name": "室外 SfM + GPS",
        "method": "多视几何重建+全球定位融合",
        "accuracy": "绝对位置 m 级；相对重建 cm～dm",
        "link": "https://colmap.github.io/",
    },
    {
        "category": "通用视觉定位",
        "name": "大场景 AprilTag 户外布局",
        "method": "扩展 Tag 布局到户外大尺度",
        "accuracy": "近距 mm；80 m 级可达 m 级容差（见论文）",
        "link": "https://www.iaarc.org/publications/fulltext/166_ISARC_2024_Paper_207.pdf",
    },
    # --- 参考对比（非纯视觉，对话中曾对比） ---
    {
        "category": "参考对比（非视觉）",
        "name": "激光跟踪仪 Leica/FARO",
        "method": "激光干涉测距+转角；单靶球跟踪",
        "accuracy": "约 0.01～0.1 mm（计量级）",
        "link": "https://www.faro.com/",
    },
    {
        "category": "参考对比（非视觉）",
        "name": "UWB 室内定位",
        "method": "超宽带 ToF/TDoA",
        "accuracy": "约 10～40 cm（NLoS 更差）",
        "link": "https://arxiv.org/html/2309.02961v2",
    },
    {
        "category": "参考对比（非视觉）",
        "name": "RTK-GPS",
        "method": "载波相位差分",
        "accuracy": "约 1～2 cm（开阔环境）",
        "link": "https://www.gps.gov/systems/gps/performance/accuracy/",
    },
    {
        "category": "参考对比（非视觉）",
        "name": "CDPR DCLM 绳长激光（非视觉）",
        "method": "绳上激光测距 vs 编码器",
        "accuracy": "绳长误差 STD 约降 58%；优于编码器数 mm",
        "link": "https://www.mdpi.com/2218-6581/10/2/60",
    },
]

HEADERS = ["类别", "名称", "方式", "精度", "链接"]


def main():
    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "全部汇总"

    ws_all.append(HEADERS)
    for row in ROWS:
        ws_all.append(
            [row["category"], row["name"], row["method"], row["accuracy"], row["link"]]
        )

    # Per-category sheets
    categories = []
    for r in ROWS:
        if r["category"] not in categories:
            categories.append(r["category"])

    for cat in categories:
        safe_title = cat.replace("/", "-")[:31]
        ws = wb.create_sheet(title=safe_title)
        ws.append(HEADERS)
        for row in ROWS:
            if row["category"] == cat:
                ws.append(
                    [row["category"], row["name"], row["method"], row["accuracy"], row["link"]]
                )

    header_font = Font(bold=True)
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"
        widths = [18, 36, 48, 36, 72]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    out = "/home/xyc/CDPR_86/src/cdpr_86_host/docs/CDPR与视觉定位精度调研.xlsx"
    wb.save(out)
    print(f"Wrote {len(ROWS)} rows to {out}")


if __name__ == "__main__":
    main()
